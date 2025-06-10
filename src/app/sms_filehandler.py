# Imports

from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.exc import IntegrityError
from sqlalchemy.sql import text

from openpyxl import load_workbook

import random as r
import string as s

from .models import StudentSMS, Student_course, Course  # Removed Hosts

# CONSTANTS
CHARACTERS = s.ascii_letters + s.digits

from . import db


# --------------------------------------------------
# Helper Functions


# Load German words from a separate file
def load_german_words():
    try:
        with open("app/german_words.txt", "r", encoding="utf-8") as file:
            words = [line.strip() for line in file if line.strip()]
            return words
    except FileNotFoundError:
        print("Error: german_words.txt not found!")
        return ["defaultword"]  # Fallback in case file is missing
    

GERMAN_WORDS = load_german_words()



def create_student(student_id, grade, grade_selector, logincode, gender=None):
    # Check if student already exists - use text query to avoid schema issues
    existing_student = db.session.execute(
        text("SELECT Student_id FROM students_sms WHERE Student_id = :student_id"),
        {"student_id": student_id}
    ).fetchone()

    if existing_student:
        print(f"Student with ID {student_id} already exists.")
        return existing_student  # Avoid duplicate entry

    try:
        # Create student without gender since column doesn't exist
        new_student = StudentSMS(
            Student_id=student_id,
            grade=grade,
            grade_selector=grade_selector,
            logincode=logincode
        )
            
        print("Student erfolgreich gespeichert!")
        db.session.add(new_student)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        print(f"IntegrityError: Duplicate ID {student_id}")
    except Exception as e:
        print(f"Error while creating a new student #{student_id}")
        print(e)



def create_course(Course_id, Course_title, Course_description, Course_teacher, Course_min_grade, Course_max_grade, Course_max_people, Course_hosts):
    try:
        new_Course = Course(
            course_id = Course_id,
            course_title = Course_title,
            course_description = Course_description,
            course_Overseers = Course_teacher,
            course_minimum_grade = Course_min_grade,
            course_maximum_grade = Course_max_grade,
            course_maximum_people = Course_max_people,
            course_hosts = Course_hosts,
        )
        db.session.add(new_Course)
        db.session.commit()
    except:
        db.session.rollback()
        print(f"Error creating a new Course #${Course_id} ${Course_title}")


def generate_login_code():
    # Select a random German word
    word = r.choice(GERMAN_WORDS)

    # Generate 5 random digits
    numbers = "".join(r.choices(s.digits, k=5))

    logincode = f"{word}{numbers}"

    try:
        # Query only the essential columns to avoid schema issues
        user_with_logincode = db.session.execute(
            text("SELECT Student_id FROM students_sms WHERE logincode = :logincode"),
            {"logincode": logincode}
        ).fetchone()

        # Ensure uniqueness
        if user_with_logincode is not None:
            return generate_login_code()

        return logincode
    except Exception as e:
        # If session is in bad state, rollback and try again
        db.session.rollback()
        print(f"Error in generate_login_code: {e}")
        return generate_login_code()


# ----------------------------------------------------------------------------
# This is the place where the magic happens ✨✨✨

def FileHandler():
    try:
        workbook = load_workbook(f"app/data/sms/uploads/workbook.xlsx")
        print("Loaded File...")

        # Clear the database more safely
        try:
            db.session.execute(text("DELETE FROM students_sms"))
            db.session.execute(text("DELETE FROM student_course"))
            db.session.execute(text("DELETE FROM courses"))
            db.session.commit()
        except Exception as e:
            print(f"Error clearing database: {e}")
            db.session.rollback()

        # Get the Students
        sheet1 = workbook.worksheets[0]

        for row_index, row in enumerate(
            sheet1.iter_rows(min_row=2, values_only=True), start=2
        ):
            student_id = row[0]  # Column A (ID)
            # Skip name columns - row[1] and row[2]
            # Skip gender column since it doesn't exist in database
            grade = row[4] if len(row) > 4 else None  # Column E (Grade)
            grade_selector = row[5] if len(row) > 5 else None  # Column F (Grade Selector)
            
            try:
                logincode = generate_login_code()
                print(f"{student_id} {grade}/{grade_selector} {logincode}")

                if student_id is None or grade is None:
                    break

                # Don't pass gender since column doesn't exist
                create_student(student_id, grade, grade_selector, logincode)
            except Exception as e:
                print(f"Error processing student {student_id}: {e}")
                db.session.rollback()
                continue

        # Get the Courses
        Course_sheet = workbook.worksheets[1]

        for row_index, row in enumerate(
            Course_sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            course_id = row[0]
            course_title = row[1]
            course_discription = row[2]
            course_teacher = row[3]
            course_min_grade = row[4]
            course_max_grade = row[5]
            course_max_people = row[6]
            course_hosts = row[7]

            print(f"{course_id} {course_title} {course_hosts} {course_discription}")

            if (
                course_id == None
                or course_title == None
                or course_discription == None
                or course_teacher == None
                or course_min_grade == None
                or course_max_grade == None
                or course_max_people == None
                or course_hosts == None
            ):
                print(f"Skipping row {row_index} due to missing data")
                break

            try:
                create_course(course_id, course_title, course_discription, course_teacher, course_min_grade, course_max_grade, course_max_people, course_hosts)
            except Exception as e:
                print(f"Error processing course {course_id}: {e}")
                db.session.rollback()
                continue

        print("Finished processing courses")
        
    except Exception as e:
        print(f"Critical error in FileHandler: {e}")
        db.session.rollback()
        raise


def FileHandlerNames():
    """Handle the names file upload separately"""
    try:
        workbook = load_workbook(f"app/data/sms/uploads/names_workbook.xlsx")
        print("Names file uploaded successfully...")
        # You can add validation or processing here if needed
        return True
    except Exception as e:
        print(f"Error processing names file: {e}")
        return False




