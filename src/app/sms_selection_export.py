from sqlalchemy.sql import text
from openpyxl import load_workbook, Workbook
import io


def load_names_map(file_storage):
    workbook = load_workbook(file_storage, keep_vba=False, data_only=True)
    sheet = workbook.worksheets[0]
    names_map = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            names_map[row[0]] = {
                'first': row[2] or "",
                'last': row[1] or "",
            }
    return names_map


def SelectionExporter(db, names_file):
    names_map = load_names_map(names_file)

    course_titles = {
        row[0]: row[1]
        for row in db.session.execute(
            text("SELECT course_id, course_title FROM courses")
        ).fetchall()
    }

    selections = db.session.execute(
        text(
            "SELECT Student_id AS student_id, Course_id AS course_id, weight "
            "FROM student_course ORDER BY Student_id ASC, weight ASC"
        )
    ).mappings().all()

    # Group courses per student, sorted by weight ascending
    student_courses = {}
    for row in selections:
        sid = row["student_id"]
        student_courses.setdefault(sid, []).append((row["weight"], row["course_id"]))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SMS_Selections"

    sheet.append(["Student ID", "Last Name", "First Name", "Wish 1", "Wish 2", "Wish 3", "Wish 4", "Wish 5", "Wish 6"])

    for student_id, courses in student_courses.items():
        names = names_map.get(student_id, {'first': '', 'last': ''})
        sorted_courses = sorted(courses, key=lambda x: x[0])
        wish_titles = [course_titles.get(cid, str(cid)) for _, cid in sorted_courses]
        # Pad to 6 columns
        while len(wish_titles) < 6:
            wish_titles.append("")
        sheet.append([student_id, names['last'], names['first']] + wish_titles[:6])

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
