#subdomain\src\app\tdw_filehandler_secure.py
# This filehandler is specifically for the Tag der Wissenschaften (TDW) functionality
# It saves NO personally identifiable information (PII) such as names or surnames of the students.
# Names are only read from Excel during export functions and kept in RAM only.

# Imports
from openpyxl import load_workbook
import random as r
import string as s
import os
from sqlalchemy import text
from sqlalchemy.exc import IntegrityError
from .models import Student, Presentation, BlockedPresentation
from . import db

# CONSTANTS
CHARACTERS = s.ascii_letters + s.digits

# --------------------------------------------------
# Helper Functions

def load_german_words():
    try:
        # Use absolute path relative to this file to be safe
        current_dir = os.path.dirname(os.path.abspath(__file__))
        word_path = os.path.join(current_dir, "german_words.txt")

        if os.path.exists(word_path):
            with open(word_path, "r", encoding="utf-8") as file:
                return [line.strip() for line in file if line.strip()]
        return ["start", "baum", "haus", "buch", "stuhl"]  # Fallback
    except Exception:
        return ["default"]

GERMAN_WORDS = load_german_words()

def generate_login_code():
    word = r.choice(GERMAN_WORDS)
    numbers = "".join(r.choices(s.digits, k=5))
    logincode = f"{word}{numbers}"

    # Check uniqueness
    exists = db.session.execute(
        db.select(Student).filter_by(logincode=logincode)
    ).scalar_one_or_none()

    if exists:
        return generate_login_code()
    return logincode

def create_student_secure(student_id, grade, grade_selector):
    """
    Creates a student with NO Name.
    """
    # Check if student already exists
    existing_student = db.session.execute(
        db.select(Student).filter_by(id=student_id)
    ).scalar_one_or_none()

    if existing_student:
        print(f"Student ID {student_id} already exists. Skipping.")
        return

    # Generate Login Code
    logincode = generate_login_code()

    try:
        new_student = Student(
            id=student_id,
            first_name=None,  # PRIVACY: Do not save name
            last_name=None,   # PRIVACY: Do not save name
            grade=grade,
            grade_selector=grade_selector,
            logincode=logincode,
        )
        db.session.add(new_student)
        db.session.commit()

    except IntegrityError:
        db.session.rollback()
        print(f"IntegrityError: Could not create student {student_id}")
    except Exception as e:
        db.session.rollback()
        print(f"Error creating student {student_id}: {e}")

def create_presentation(presentation_id, title, presenter, abstract, grades):
    try:
        new_presentation = Presentation(
            id=presentation_id,
            title=title,
            presenter=presenter,
            abstract=abstract,
            grades=grades,
        )
        db.session.add(new_presentation)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error creating presentation {presentation_id}: {e}")

def add_blocked_presentation(student_id, presentation_id):
    try:
        new_blocked_presentation = BlockedPresentation(
            student_id=student_id,
            presentation_id=presentation_id
        )
        db.session.add(new_blocked_presentation)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error creating blocked presentation {student_id}/{presentation_id}: {e}")

# ----------------------------------------------------------------------------
# Secure File Handler

def FileHandlerTDWSecure():
    """
    Reads workbook.xlsx but IGNORES name columns.
    Only saves ID, Grade, Class.
    """
    try:
        file_path = os.path.join("app", "data", "tdw", "uploads", "workbook.xlsx")
        workbook = load_workbook(file_path)
        print("Loaded Workbook (Secure Mode)...")

        # Clear Database
        try:
            db.session.execute(text("DELETE FROM students"))
            db.session.execute(text("DELETE FROM presentations"))
            db.session.execute(text("DELETE FROM selections"))
            db.session.execute(text("DELETE FROM blocked_presentations"))
            db.session.commit()
            print("Database cleared.")
        except Exception as e:
            print(f"Error clearing database: {e}")
            db.session.rollback()

        # 1. Process Students (Sheet 1)
        sheet1 = workbook.worksheets[0]

        for row in sheet1.iter_rows(min_row=2, values_only=True):
            student_id = row[0]  # Column A: ID

            # SKIPPING Columns B, C (Last Name, First Name)

            grade = row[4]           # Column E: Grade
            grade_selector = row[5]  # Column F: Class

            if student_id is None or grade is None:
                continue

            create_student_secure(student_id, grade, grade_selector)

        # 2. Process Presentations (Sheet 3 - index 2)
        if len(workbook.worksheets) > 2:
            presentations_sheet = workbook.worksheets[2]

            for row in presentations_sheet.iter_rows(min_row=2, values_only=True):
                presentation_id = row[0]
                title = row[1]
                presenter = row[2]
                abstract = row[3]

                if presentation_id is None or title is None:
                    continue

                # Process grades
                grades = []
                g = 5
                for grade in row[4:12]:
                    if grade == -1:
                        grades.append(g)
                    g += 1

                grades = str(grades)[1:-1]

                create_presentation(presentation_id, title, presenter, abstract, grades)

        # 3. Process Blocked Presentations (Sheet 4 - index 3)
        if len(workbook.worksheets) > 3:
            blocked_presentations_sheet = workbook.worksheets[3]

            for row in blocked_presentations_sheet.iter_rows(min_row=2, values_only=True):
                student_id = row[0]
                presentation_id = row[1]

                if student_id is None or presentation_id is None:
                    continue

                add_blocked_presentation(student_id, presentation_id)

        print("Secure Data processing complete. No PII saved.")

    except Exception as e:
        print(f"Critical error in FileHandlerTDWSecure: {e}")
        db.session.rollback()

def load_names_map(file_storage):
    """
    Helper for Secure Exports:
    Reads names from uploaded file into RAM dictionary.
    """
    workbook = load_workbook(file_storage)
    sheet = workbook.worksheets[0]
    names_map = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            names_map[row[0]] = {
                'last': row[1],
                'first': row[2]
            }
    return names_map
