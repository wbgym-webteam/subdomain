from sqlalchemy.sql import text
from openpyxl import load_workbook, Workbook
import io


def load_names_map(file_storage):
    workbook = load_workbook(file_storage, keep_vba=False, data_only=True)
    sheet = workbook.worksheets[0]
    names_map = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            names_map[row[0]] = {
                'first': row[2] or "",
                'last': row[1] or "",
            }
    return names_map


def AssignmentExporter(db, names_file):
    names_map = load_names_map(names_file)

    students = db.session.execute(
        text("SELECT Student_id, grade, grade_selector FROM students_sms ORDER BY grade, grade_selector, Student_id")
    ).mappings().all()

    assignments = {}
    for row in db.session.execute(
        text("SELECT student_id, course_id, session FROM sms_assignment")
    ).mappings().all():
        assignments[(row["student_id"], row["session"])] = row["course_id"]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Kurs_Zuteilungen"
    sheet.append(["Student ID", "Last Name", "First Name", "Grade", "Grade Selector",
                  "Session 1 Course", "Session 2 Course"])

    for s in students:
        sid = s["Student_id"]
        names = names_map.get(sid, {"first": "", "last": ""})
        c1 = assignments.get((sid, 1)) or ""
        c2 = assignments.get((sid, 2)) or ""
        sheet.append([
            sid, names["last"], names["first"],
            s["grade"], s["grade_selector"],
            c1, c2
        ])

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
