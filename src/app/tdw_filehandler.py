# Imports

from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.exc import IntegrityError

from openpyxl import load_workbook

import random as r
import string as s

from .models import Presentation, Student

# CONSTANTS
CHARACTERS = s.ascii_letters + s.digits

from . import db

# --------------------------------------------------
# Helper Functions


# Load German words from a separate file
def load_german_words():
    try:
        with open("app/german_words.txt", "r", encoding="utf-8") as file:
            words = [line.strip() for line in file if line.strip()]
            return words
    except FileNotFoundError:
        print("Error: german_words.txt not found!")
        return ["defaultword"]  # Fallback in case file is missing


GERMAN_WORDS = load_german_words()


def create_student(student_id, last_name, first_name, grade, grade_selector, logincode):
    # Check if student already exists
    existing_student = db.session.execute(
        db.select(Student).filter_by(id=student_id)
    ).scalar_one_or_none()

    if existing_student:
        print(existing_student.last_name)
        print(f"Student with ID {student_id} already exists.")
        return existing_student  # Avoid duplicate entry

    try:
        new_student = Student(
            id=student_id,
            last_name=last_name,
            first_name=first_name,
            grade=grade,
            grade_selector=grade_selector,
            logincode=logincode,
        )
        print("Student erfolgreich gespeichert!")
        db.session.add(new_student)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        print(f"IntegrityError: Duplicate ID {student_id}")
    except Exception as e:
        print(
            f"Error while creating a new student #${student_id} {last_name} {first_name}"
        )
        print(e)


def generate_login_code():

    # Select a random German word
    word = r.choice(GERMAN_WORDS)

    # Generate 5 random digits
    numbers = "".join(r.choices(s.digits, k=5))

    logincode = f"{word}{numbers}"

    user_with_logincode = db.session.execute(
        db.select(Student).filter_by(logincode=logincode)
    ).scalar_one_or_none()

    # Ensure uniqueness
    if user_with_logincode is not None:
        return generate_login_code()

    # print(Student.query().filter_by(logincode=logincode).first())

    return logincode

    # if mine is wrong or does not work, use this one

    # logincode = "".join(r.choice(CHARACTERS) for _ in range(8))

    # if logincode_exists(logincode):
    # return generateLoginCode()
    # else:
    # return logincode


def create_presentation(presentation_id, title, presenter, abstract, grades):
    try:
        new_presentation = Presentation(
            id=presentation_id,
            title=title,
            presenter=presenter,
            abstract=abstract,
            grades=grades,
        )
        db.session.add(new_presentation)
        db.session.commit()
    except:
        db.session.rollback()
        print(f"Error creating a new presentation #${presentation_id} ${title}")


# ------------------------------------------------------------------------------
# This is the place where the magic happens ✨✨✨


def FileHandler():
    workbook = load_workbook(f"app/data/tdw/uploads/workbook.xlsx")
    print("Loaded File...")

    # Get the Students
    sheet1 = workbook.worksheets[0]

    for row_index, row in enumerate(
        sheet1.iter_rows(min_row=2, values_only=True), start=2
    ):
        student_id = row[0]  # Column A (ID)
        last_name = row[1]  # Column B (Last Name)
        first_name = row[2]  # Column C (First Name)
        grade = row[4]  # Column E (Grade)
        grade_selector = row[5]  # Column F (Grade Selector)
        logincode = generate_login_code()

        print(
            f"{student_id} {last_name} {first_name} {grade}/{grade_selector} {logincode}"
        )

        if (
            student_id == None
            or last_name == None
            or first_name == None
            or grade == None
        ):
            break

        create_student(
            student_id, last_name, first_name, grade, grade_selector, logincode
        )

    # Get the Presentations
    presentations_sheet = workbook.worksheets[2]

    for row_index, row in enumerate(
        presentations_sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        presentation_id = row[0]
        title = row[1]
        presenter = row[2]
        abstract = row[3]

        grades = []
        g = 5
        for grade in row[4:12]:
            if grade == -1:
                grades.append(g)
            else:
                pass
            g += 1

        grades = str(grades)[1:-1]

        print(f"{presentation_id} {title} {presenter} {abstract} {grades}")

        if (
            presentation_id == None
            or title == None
            or presenter == None
            or abstract == None
        ):
            break

        create_presentation(presentation_id, title, presenter, abstract, grades)
