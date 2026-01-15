# Imports
# this filehandler is specifically for the project week (PT) functionality
# It saved additionaly the names, surnames and emails of the students, which is currently not done in the current settings but may be changed at a later point
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.exc import IntegrityError
from sqlalchemy.sql import text


from openpyxl import load_workbook

import random as r
import string as s
import os

from .models import PTStudent, PTPresentation

# CONSTANTS
CHARACTERS = s.ascii_letters + s.digits

from . import db

# --------------------------------------------------
# Helper Functions

# Load German words from a separate file
def load_german_words():
    try:
        with open("app/german_words.txt", "r", encoding="utf-8") as file:
            words = [line.strip() for line in file if line.strip()]
            return words
    except FileNotFoundError:
        print("Error: german_words.txt not found!")
        return ["defaultword"]  # Fallback in case file is missing
    
GERMAN_WORDS = load_german_words()

def load_names_map(file_storage):
    """
    Reads an uploaded Excel file in memory and returns a dictionary.
    Format: { student_id_int: {'first': 'John', 'last': 'Doe'} }
    """
    # Use keep_vba=False and data_only=True for .xlsm files (macro-enabled)
    workbook = load_workbook(file_storage, keep_vba=False, data_only=True)
    sheet = workbook.active
    
    names_map = {}
    
    # Assuming Row 1 is headers, data starts Row 2
    # Column A=ID, B=Last Name, C=First Name (Adjust indices based on your excel)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        s_id = row[0] 
        last_name = row[1]
        first_name = row[2]
        
        if s_id is not None:
            names_map[s_id] = {
                'first': first_name,
                'last': last_name
            }
            
    return names_map

def create_student(student_id, last_name, first_name, email, grade, grade_selector, logincode, gender):
    # Check if student already exists
    existing_student = db.session.execute(
        db.select(PTStudent).filter_by(id=student_id)
    ).scalar_one_or_none()

    if existing_student:
        print(existing_student.last_name)
        print(f"Student with ID {student_id} already exists.")
        return existing_student  # Avoid duplicate entry

    try:
        new_student = PTStudent(
            id=student_id,
            last_name=last_name,
            first_name=first_name,
            email=email if email else f"student{student_id}@example.com",  # Provide default email
            password_hash="dummy_hash",  # Provide dummy hash since we use logincode
            grade=grade,
            grade_selector=grade_selector,
            logincode=logincode,
            gender=gender,
        )
        print("Student erfolgreich gespeichert!")
        db.session.add(new_student)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        print(f"IntegrityError: Duplicate ID {student_id}")
    except Exception as e:
        print(
            f"Error while creating a new student #${student_id} {last_name} {first_name}"
        )
        print(e)


def generate_login_code():

    # Select a random German word
    word = r.choice(GERMAN_WORDS)

    # Generate 5 random digits
    numbers = "".join(r.choices(s.digits, k=5))

    logincode = f"{word}{numbers}"

    user_with_logincode = db.session.execute(
        db.select(PTStudent).filter_by(logincode=logincode)
    ).scalar_one_or_none()

    # Ensure uniqueness
    if user_with_logincode is not None:
        return generate_login_code()

    # print(Student.query().filter_by(logincode=logincode).first())

    return logincode

    # if mine is wrong or does not work, use this one

    # logincode = "".join(r.choice(CHARACTERS) for _ in range(8))

    # if logincode_exists(logincode):
    # return generateLoginCode()
    # else:
    # return logincode



def create_pt_presentation(id, title, description, presenter, slot, max_people, column, room, gender):
    try:
        new_pt_presentation = PTPresentation(
            id=id,
            title=title,
            description=description,
            presenter=presenter,     
            slot=slot,             
            max_students=max_people,
            column=column,     
            room=room,              
            gender=gender
        )
        db.session.add(new_pt_presentation)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error creating a new pt_presentation #{id} {title}: {e}")
        
# ----------------------------------------------------------------------------
# This is the place where the magic happens ✨✨✨

def FileHandlerPT():
    try:
        # Use absolute path
        current_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(current_dir, "data", "pt", "uploads", "workbook.xlsx")
        # Use keep_vba=False and data_only=True for .xlsm files (macro-enabled)
        workbook = load_workbook(file_path, keep_vba=False, data_only=True)
        print("Loaded File...")

        # Clear the database more safely
        try:
            db.session.execute(text("DELETE FROM pt_students"))
            db.session.execute(text("DELETE FROM pt_selections"))
            db.session.execute(text("DELETE FROM pt_presentations"))
            db.session.commit()
        except Exception as e:
            print(f"Error clearing database: {e}")
            db.session.rollback()

        # Get the Students
        sheet1 = workbook.worksheets[0]

        for row_index, row in enumerate(
            sheet1.iter_rows(min_row=2, values_only=True), start=2
        ):
            student_id = row[0]  # Column A (ID)
            last_name = row[1] if len(row) > 1 else "Unknown"  # Column B
            first_name = row[2] if len(row) > 2 else "Unknown"  # Column C
            email = row[3] if len(row) > 3 else None  # Column D
            grade = row[4] if len(row) > 4 else None  # Column E (Grade)
            grade_selector = row[5] if len(row) > 5 else None  # Column F (Grade Selector)
            gender = row[6] if len(row) > 6 else "u"  # Column G
            
            try:
                logincode = generate_login_code()
                # print(f"{student_id} {last_name} {first_name} {grade}/{grade_selector} {logincode}")

                if student_id is None or grade is None:
                    break

                # Pass all required parameters
                create_student(student_id, last_name, first_name, email, grade, grade_selector, logincode, gender)
            except Exception as e:
                print(f"Error processing student {student_id}: {e}")
                db.session.rollback()
                continue

        # Get the Courses
        Course_sheet = workbook.worksheets[1]

        for row_index, row in enumerate(
            Course_sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            PTPresentation_id = row[0]
            PTPresentation_title = row[1]
            PTPresentation_hosts = row[2]
            PTPresentation_teacher = row[3]
            PTPresentation_description = row[4]
            PTPresentation_slot = row[5]
            PTPresentation_max_people = row[6]
            PTPresentation_column = row[7]
            PTPresentation_room = row[8]  
            PTPresenation_gender = row[9]
            # Add column data from Excel

            # print(f"{PTPresentation_id} {PTPresentation_title} {PTPresentation_hosts} {PTPresentation_description} Column: {PTPresentation_column}")

            if (
                PTPresentation_id == None
                or PTPresentation_title == None
            ):
                print(f"Skipping row {row_index} due to missing data")
                break

            try:
                # Using Keyword Arguments to ensure correct mapping
                create_pt_presentation(
                    id=PTPresentation_id,
                    title=PTPresentation_title,
                    description=PTPresentation_description,
                    presenter=PTPresentation_hosts,  
                    slot=PTPresentation_slot,    
                    max_people=PTPresentation_max_people,
                    column=PTPresentation_column,
                    room=PTPresentation_room,    
                    gender=PTPresenation_gender
                )
            except Exception as e:
                print(f"Error processing course {PTPresentation_id}: {e}")
                db.session.rollback()
                continue

        print("Finished processing courses")
        
    except Exception as e:
        print(f"Critical error in FileHandler: {e}")
        db.session.rollback()
        raise


def FileHandlerNames():
    """Handle the names file upload separately"""
    try:
        # Use absolute path
        current_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(current_dir, "data", "pt", "uploads", "names_workbook.xlsx")
        # Use keep_vba=False and data_only=True for .xlsm files (macro-enabled)
        workbook = load_workbook(file_path, keep_vba=False, data_only=True)
        print("Names file uploaded successfully...")
        # You can add validation or processing here if needed
        return True
    except Exception as e:
        print(f"Error processing names file: {e}")
        return False