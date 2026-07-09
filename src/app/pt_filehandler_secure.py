#subdomain\src\app\pt_filehandler_secure.py
# This filehandler is specifically for the project week (PT) functionality
# It saves NO personally identifiable information (PII) such as names, surnames, or emails of the students.
# this version is currently used, but may be scrapped later, if we may be allowed to store Personal Data on the server

# Impots
from openpyxl import load_workbook
import random as r
import string as s
import os
from sqlalchemy import text
from sqlalchemy.exc import IntegrityError
from .models import PTStudent, PTPresentation
from . import db

# CONSTANTS
CHARACTERS = s.ascii_letters + s.digits

# --------------------------------------------------
# Helper Functions

def load_german_words():
    try:
        # Use absolute path relative to this file to be safe
        current_dir = os.path.dirname(os.path.abspath(__file__))
        # Assuming german_words.txt is in 'app/' folder (one level up if this is in app/utils, 
        # but since this is in app/, it's just adjacent or in data)
        # Adjust path if needed based on your structure
        word_path = os.path.join(current_dir, "german_words.txt")
        
        if os.path.exists(word_path):
             with open(word_path, "r", encoding="utf-8") as file:
                return [line.strip() for line in file if line.strip()]
        return ["start", "baum", "haus", "buch", "stuhl"] # Fallback
    except Exception:
        return ["default"]

GERMAN_WORDS = load_german_words()

def generate_login_code():
    word = r.choice(GERMAN_WORDS)
    numbers = "".join(r.choices(s.digits, k=5))
    logincode = f"{word}{numbers}"
    
    # Check uniqueness
    exists = db.session.execute(
        db.select(PTStudent).filter_by(logincode=logincode)
    ).scalar_one_or_none()
    
    if exists:
        return generate_login_code()
    return logincode

def create_student_secure(student_id, grade, grade_selector, gender):
    """
    Creates a student with NO Name and a Dummy Email.
    """
    # Check if student already exists
    existing_student = db.session.execute(
        db.select(PTStudent).filter_by(id=student_id)
    ).scalar_one_or_none()

    if existing_student:
        print(f"Student ID {student_id} already exists. Skipping.")
        return

    # Generate a Dummy Email to satisfy DB 'Unique' constraints
    dummy_email = f"{student_id}_anon@pt.system"
    
    # Generate Login Code
    logincode = generate_login_code()

    try:
        new_student = PTStudent(
            id=student_id,
            first_name=None,  # PRIVACY: Do not save name
            last_name=None,   # PRIVACY: Do not save name
            email=None, # PRIVACY: Do not save email
            password_hash="dummy_hash",
            grade=grade,
            grade_selector=grade_selector,
            logincode=logincode,
            gender=gender,
        )
        db.session.add(new_student)
        db.session.commit()
        
    except IntegrityError:
        db.session.rollback()
        print(f"IntegrityError: Could not create student {student_id}")
    except Exception as e:
        db.session.rollback()
        print(f"Error creating student {student_id}: {e}")

def create_pt_presentation(id, title, description, presenter, teacher, slot, max_people, column, room, gender):
    try:
        new_pt_presentation = PTPresentation(
            id=id,
            title=title,
            description=description,
            presenter=presenter,  
            teacher=teacher,   
            slot=slot,             
            max_students=max_people,
            column=column,     
            room=room,              
            gender=gender
        )
        db.session.add(new_pt_presentation)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error creating presentation {id}: {e}")

# ----------------------------------------------------------------------------
# Secure File Handler

def FileHandlerPTSecure():
    """
    Reads workbook.xlsx but IGNORES name/email columns.
    Only saves ID, Grade, Class, Gender.
    """
    # Use absolute path
    current_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(current_dir, "data", "pt", "uploads", "workbook.xlsx")

    # Check if file exists
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Workbook file not found at {file_path}")

    try:
        # Use keep_vba=False and data_only=True for .xlsm files (macro-enabled)
        # This prevents issues with VBA content and formula evaluation
        workbook = load_workbook(file_path, keep_vba=False, data_only=True)
        print("Loaded Workbook (Secure Mode)...")

        # Clear Database
        try:
            db.session.execute(text("DELETE FROM pt_students"))
            db.session.execute(text("DELETE FROM pt_selections"))
            db.session.execute(text("DELETE FROM pt_presentations"))
            db.session.execute(text("DELETE FROM pt_assignments"))
            db.session.commit()
            print("Database cleared.")
        except Exception as e:
            print(f"Error clearing database: {e}")
            db.session.rollback()
            raise Exception(f"Failed to clear database: {e}")

        # 1. Process Students (Sheet 1)
        if len(workbook.worksheets) < 1:
            raise ValueError("Workbook must have at least one sheet for students")

        sheet1 = workbook.worksheets[0]

        for row in sheet1.iter_rows(min_row=2, values_only=True):
            student_id = row[0] # Column A: ID

            # SKIPPING Columns B, C, D (Name, Surname, Email)

            grade = row[4] if len(row) > 4 else None           # Column E: Grade
            grade_selector = row[5] if len(row) > 5 else None  # Column F: Class
            gender = row[6] if len(row) > 6 else "u" # Column G: Gender

            if student_id is None or grade is None:
                continue

            create_student_secure(student_id, grade, grade_selector, gender)

        # 2. Process Courses (Sheet 2)
        if len(workbook.worksheets) > 1:
            course_sheet = workbook.worksheets[1]
            for row in course_sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is None: continue

                create_pt_presentation(
                    id=row[0],
                    title=row[1],
                    presenter=row[2],
                    teacher=row[3],
                    description=row[4],
                    slot=row[5],
                    max_people=row[6],
                    column=row[7],
                    room=row[8],
                    gender=row[9] if len(row) > 9 else 'u'
                )

        print("Secure Data processing complete. No PII saved.")

    except Exception as e:
        print(f"Critical error in FileHandlerPTSecure: {e}")
        db.session.rollback()
        raise  # Re-raise the exception so the route handler can catch it

def load_names_map(file_storage):
    """
    Helper for Secure Exports:
    Reads names from uploaded file into RAM dictionary.
    """
    # Use keep_vba=False and data_only=True for .xlsm files (macro-enabled)
    workbook = load_workbook(file_storage, keep_vba=False, data_only=True)
    sheet = workbook.worksheets[0]
    names_map = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            names_map[row[0]] = {
                'last': row[1],
                'first': row[2]
            }
    return names_map